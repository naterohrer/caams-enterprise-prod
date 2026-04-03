"""XLSX export — Coverage Report, Evidence Checklist, SOA, Findings, Recommendations."""

import io
from datetime import datetime, timezone

from fastapi import APIRouter, Depends, HTTPException, Request
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy.orm import Session

from app import models
from app.auth import require_viewer
from app.database import get_db
from app.engine import mapper
from app.limiter import limiter
from app.routers.assessments import _findings_map, _notes_map, _ownership_map

router = APIRouter(prefix="/assessments", tags=["export"])

# Colours
GREEN = PatternFill("solid", fgColor="C6EFCE")
AMBER = PatternFill("solid", fgColor="FFEB9C")
RED = PatternFill("solid", fgColor="FFC7CE")
GREY = PatternFill("solid", fgColor="D9D9D9")
BLUE_HEADER = PatternFill("solid", fgColor="1F3864")
BOLD_WHITE = Font(bold=True, color="FFFFFF")


def _color_for(status: str):
    return {"covered": GREEN, "partial": AMBER, "not_covered": RED,
            "not_applicable": GREY}.get(status, GREY)


def _col_widths(ws, widths: list[int]):
    from openpyxl.utils import get_column_letter
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _header_row(ws, headers: list[str], row: int = 1):
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.font = BOLD_WHITE
        cell.fill = BLUE_HEADER
        cell.alignment = Alignment(wrap_text=True)


@router.get("/{assessment_id}/export")
@limiter.limit("10/minute")
def export_xlsx(
    request: Request,
    assessment_id: int,
    _: models.User = Depends(require_viewer),
    db: Session = Depends(get_db),
):
    a = db.query(models.Assessment).filter(models.Assessment.id == assessment_id).first()
    if not a:
        raise HTTPException(status_code=404, detail="Assessment not found")

    controls = a.framework.controls if a.framework else []
    nm = _notes_map(assessment_id, db)
    om = _ownership_map(assessment_id, db)
    fm = _findings_map(assessment_id, db)
    cov = mapper.compute_coverage(controls, a.tools, nm, om, fm)
    all_tools = db.query(models.Tool).all()
    recs = mapper.compute_recommendations(controls, a.tools, all_tools)

    wb = Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    rows = [
        ("Assessment", a.name),
        ("Framework", a.framework.name if a.framework else ""),
        ("Status", a.status.title()),
        ("Score", f"{cov['score']}%"),
        ("Covered", cov["covered"]),
        ("Partial", cov["partial"]),
        ("Not Covered", cov["not_covered"]),
        ("Not Applicable", cov["not_applicable"]),
        ("Total Controls", cov["total_controls"]),
        ("Tool Count", len(a.tools)),
        ("Exported", datetime.now(timezone.utc).replace(tzinfo=None).strftime("%Y-%m-%d %H:%M UTC")),
    ]
    for r, (k, v) in enumerate(rows, 1):
        ws.cell(row=r, column=1, value=k).font = Font(bold=True)
        ws.cell(row=r, column=2, value=str(v))
    _col_widths(ws, [25, 40])

    # ── Sheet 2: Coverage Report ───────────────────────────────────────────────
    ws2 = wb.create_sheet("Coverage Report")
    headers = ["Control ID", "Title", "Status", "Override", "Owner", "Team",
               "Evidence Owner", "Review Status", "Assignee", "Due Date",
               "Overdue", "Covered By", "Missing Tags", "Notes", "Evidence URL",
               "Override Justification", "Finding Count", "Open Findings"]
    _header_row(ws2, headers)
    for row_num, c in enumerate(cov["controls"], 2):
        due = c["due_date"].strftime("%Y-%m-%d") if c["due_date"] else ""
        values = [
            c["control_id"], c["title"], c["status"].replace("_", " ").title(),
            "Yes" if c["is_override"] else "", c["owner"], c["team"],
            c["evidence_owner"], c["review_status"], c["assignee"], due,
            "Yes" if c["is_overdue"] else "",
            ", ".join(c["covered_by"]), ", ".join(c["missing_tags"]),
            c["notes"], c["evidence_url"], c["override_justification"],
            c["finding_count"], c["open_finding_count"],
        ]
        for col, v in enumerate(values, 1):
            cell = ws2.cell(row=row_num, column=col, value=v)
            if col == 3:
                cell.fill = _color_for(c["status"])
        ws2.row_dimensions[row_num].height = 30
    _col_widths(ws2, [12, 35, 14, 10, 20, 20, 20, 14, 20, 12,
                      10, 30, 30, 40, 40, 40, 14, 14])

    # ── Sheet 3: Evidence Checklist ────────────────────────────────────────────
    ws3 = wb.create_sheet("Evidence Checklist")
    _header_row(ws3, ["Control ID", "Title", "Evidence Item", "Owner", "Evidence URL", "Status", "Notes"])
    row_num = 2
    for c in cov["controls"]:
        if not c["is_applicable"]:
            continue
        for item in c["evidence_items"]:
            values = [c["control_id"], c["title"], item, c["evidence_owner"],
                      c["evidence_url"], c["status"].replace("_", " ").title(), c["notes"]]
            for col, v in enumerate(values, 1):
                cell = ws3.cell(row=row_num, column=col, value=v)
                if col == 6:
                    cell.fill = _color_for(c["status"])
            row_num += 1
    _col_widths(ws3, [12, 30, 50, 20, 40, 14, 40])

    # ── Sheet 4: Statement of Applicability (SOA) ─────────────────────────────
    ws4 = wb.create_sheet("SOA")
    _header_row(ws4, ["Control ID", "Title", "Applicable", "Exclusion Reason",
                      "Status", "Owner", "Override", "Override Justification",
                      "Review Status", "Evidence URL"])
    for row_num, c in enumerate(cov["controls"], 2):
        values = [
            c["control_id"], c["title"],
            "Yes" if c["is_applicable"] else "No",
            c["exclusion_reason"],
            c["status"].replace("_", " ").title() if c["is_applicable"] else "N/A",
            c["owner"],
            "Yes" if c["is_override"] else "",
            c["override_justification"],
            c["review_status"],
            c["evidence_url"],
        ]
        for col, v in enumerate(values, 1):
            cell = ws4.cell(row=row_num, column=col, value=v)
            if col == 5:
                cell.fill = _color_for(c["status"])
    _col_widths(ws4, [12, 35, 12, 40, 14, 20, 10, 40, 14, 40])

    # ── Sheet 5: Findings ────────────────────────────────────────────────────
    findings = db.query(models.Finding).filter(
        models.Finding.assessment_id == assessment_id
    ).order_by(models.Finding.severity, models.Finding.created_at).all()
    ws5 = wb.create_sheet("Findings")
    _header_row(ws5, ["ID", "Control ID", "Title", "Severity", "Status",
                      "Remediation Owner", "Target Date", "Close Date", "Notes"])
    sev_color = {"critical": RED, "high": RED, "medium": AMBER,
                 "low": GREEN, "informational": GREY}
    for row_num, f in enumerate(findings, 2):
        td = f.target_date.strftime("%Y-%m-%d") if f.target_date else ""
        cd = f.actual_close_date.strftime("%Y-%m-%d") if f.actual_close_date else ""
        values = [f.id, f.control_id, f.title, f.severity, f.status,
                  f.remediation_owner, td, cd, f.notes]
        for col, v in enumerate(values, 1):
            cell = ws5.cell(row=row_num, column=col, value=v)
            if col == 4:
                cell.fill = sev_color.get(f.severity, GREY)
    _col_widths(ws5, [8, 12, 40, 12, 14, 25, 12, 12, 40])

    # ── Sheet 6: Recommendations ──────────────────────────────────────────────
    if recs:
        ws6 = wb.create_sheet("Recommendations")
        _header_row(ws6, ["Tool", "Category", "Controls Helped", "Gaps Closed"])
        for row_num, r in enumerate(recs, 2):
            ws6.cell(row=row_num, column=1, value=r["tool_name"])
            ws6.cell(row=row_num, column=2, value=r["category"])
            ws6.cell(row=row_num, column=3, value=r["controls_helped"])
            ws6.cell(row=row_num, column=4, value=", ".join(r["gaps_closed"]))
        _col_widths(ws6, [30, 20, 18, 60])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"caams_{a.name.replace(' ', '_')}_{datetime.now(timezone.utc).replace(tzinfo=None).strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/{assessment_id}/export/soa")
@limiter.limit("10/minute")
def export_soa(
    request: Request,
    assessment_id: int,
    _: models.User = Depends(require_viewer),
    db: Session = Depends(get_db),
):
    """Standalone SOA export (PDF-ready XLSX)."""
    a = db.query(models.Assessment).filter(models.Assessment.id == assessment_id).first()
    if not a:
        raise HTTPException(status_code=404, detail="Assessment not found")

    controls = a.framework.controls if a.framework else []
    nm = _notes_map(assessment_id, db)
    om = _ownership_map(assessment_id, db)
    fm = _findings_map(assessment_id, db)
    cov = mapper.compute_coverage(controls, a.tools, nm, om, fm)

    # Bulk-load all reviewers referenced by notes to avoid N+1 queries
    reviewer_ids = {n.reviewed_by_id for n in nm.values() if n.reviewed_by_id}
    reviewer_map: dict = {}
    if reviewer_ids:
        users = db.query(models.User).filter(models.User.id.in_(reviewer_ids)).all()
        reviewer_map = {u.id: u.username for u in users}

    wb = Workbook()
    ws = wb.active
    ws.title = "Statement of Applicability"

    # Title block
    ws["A1"] = "Statement of Applicability"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A2"] = f"Assessment: {a.name}"
    ws["A3"] = f"Framework: {a.framework.name if a.framework else 'N/A'} {a.framework.version if a.framework else ''}"
    ws["A4"] = f"Status: {a.status.title()}"
    ws["A5"] = f"Generated: {datetime.now(timezone.utc).replace(tzinfo=None).strftime('%Y-%m-%d %H:%M UTC')}"

    _header_row(ws, ["Control ID", "Control Title", "Applicable", "Exclusion Reason",
                      "Implementation Status", "Owner", "Team",
                      "Manual Override", "Override Justification", "Override Expires",
                      "Review Status", "Reviewer", "Evidence URL",
                      "Required Evidence Items"], row=6)

    for row_num, c in enumerate(cov["controls"], 7):
        note = nm.get(c["control_id"])
        reviewer_name = reviewer_map.get(note.reviewed_by_id, "") if note and note.reviewed_by_id else ""

        exp = ""
        if note and note.override_expires:
            exp = note.override_expires.strftime("%Y-%m-%d")

        values = [
            c["control_id"],
            c["title"],
            "Yes" if c["is_applicable"] else "No",
            c["exclusion_reason"],
            c["status"].replace("_", " ").title() if c["is_applicable"] else "Not Applicable",
            c["owner"],
            c["team"],
            "Yes" if c["is_override"] else "",
            c["override_justification"],
            exp,
            c["review_status"],
            reviewer_name,
            c["evidence_url"],
            "; ".join(c["evidence_items"]),
        ]
        for col, v in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col, value=v)
            if col == 5:
                cell.fill = _color_for(c["status"])

    _col_widths(ws, [12, 35, 12, 35, 18, 20, 20, 14, 40, 14, 14, 20, 40, 60])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"SOA_{a.name.replace(' ', '_')}_{datetime.now(timezone.utc).replace(tzinfo=None).strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
