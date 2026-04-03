"""Import CIS Controls from official XLSX workbook."""

import io
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app import models


def parse_cis_xlsx(
    xlsx_path: Path,
    version: str = "v8",
    sheet_name: Optional[str] = None,
    min_ig: int = 1,
) -> dict:
    """Parse a CIS Controls XLSX workbook into a CAAMS framework dict.

    Returns a dict with keys: name, version, controls.
    Each control has: control_id, title, description, required_tags,
    optional_tags, evidence, sub_controls.

    Args:
        xlsx_path: Path to the XLSX file.
        version: Version label to embed in the framework dict (e.g. "v8.1").
        sheet_name: Worksheet name to parse; auto-detected if None.
        min_ig: Minimum Implementation Group level to include (1=all, 2=IG1+2).
    """
    wb = load_workbook(Path(xlsx_path), read_only=True)
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"
            )
        ws = wb[sheet_name]
    else:
        ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"name": "CIS Controls", "version": version, "controls": []}

    # Detect header row by looking for recognisable column names in the first 5 rows
    header_row_idx = 0
    headers: list = []
    for i, row in enumerate(rows[:5]):
        if row and any(
            str(c or "").strip().lower()
            in ("control id", "safeguard", "ig1", "title", "#")
            for c in row
        ):
            header_row_idx = i
            headers = [str(c or "").strip().lower() for c in row]
            break

    def _col(*candidates: str) -> int:
        for name in candidates:
            if name in headers:
                return headers.index(name)
        return -1

    id_col = _col("control id", "safeguard", "cis control", "id", "#")
    title_col = _col("title", "safeguard title", "control title", "name")
    desc_col = _col("description", "why it matters", "overview")
    ig1_col = _col("ig1", "ig 1")
    ig2_col = _col("ig2", "ig 2")
    ig3_col = _col("ig3", "ig 3")

    # Fall back to positional indices if headers were not detected
    if id_col == -1:
        id_col = 0
    if title_col == -1:
        title_col = 1
    if desc_col == -1:
        desc_col = 2

    data_rows = rows[header_row_idx + 1 :] if headers else rows[1:]

    def _cell(row: tuple, idx: int) -> str:
        return str(row[idx]).strip() if idx < len(row) and row[idx] else ""

    def _row_ig(row: tuple) -> int:
        if ig3_col != -1 and ig3_col < len(row) and row[ig3_col]:
            return 3
        if ig2_col != -1 and ig2_col < len(row) and row[ig2_col]:
            return 2
        return 1

    controls_map: dict = {}
    controls_order: list = []

    for row in data_rows:
        if not row or not row[id_col]:
            continue
        cid = str(row[id_col]).strip()
        if not cid or cid.lower() in ("control id", "safeguard", "n/a", ""):
            continue

        title = _cell(row, title_col)
        description = _cell(row, desc_col)

        # IG filtering: skip rows whose IG level exceeds the requested minimum
        if min_ig < 3 and ig1_col != -1:
            if _row_ig(row) > min_ig:
                continue

        is_sub = "." in cid
        if is_sub:
            parent_id = cid.split(".")[0]
            if parent_id not in controls_map:
                controls_map[parent_id] = {
                    "control_id": parent_id,
                    "title": f"CIS Control {parent_id}",
                    "description": "",
                    "required_tags": [],
                    "optional_tags": [],
                    "evidence": [],
                    "sub_controls": [],
                }
                controls_order.append(parent_id)
            controls_map[parent_id]["sub_controls"].append(
                {"id": cid, "title": title, "description": description}
            )
        else:
            if cid not in controls_map:
                controls_map[cid] = {
                    "control_id": cid,
                    "title": title,
                    "description": description,
                    "required_tags": [],
                    "optional_tags": [],
                    "evidence": [],
                    "sub_controls": [],
                }
                controls_order.append(cid)
            else:
                # Back-fill title/description on an auto-created parent stub
                if title:
                    controls_map[cid]["title"] = title
                if description:
                    controls_map[cid]["description"] = description

    return {
        "name": "CIS Controls",
        "version": version,
        "controls": [controls_map[pid] for pid in controls_order],
    }


def import_cis_xlsx(content: bytes, db: Session) -> dict:
    wb = load_workbook(io.BytesIO(content), read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    if not rows:
        return {"imported": 0, "skipped": 0, "message": "No data rows found"}

    fw = db.query(models.Framework).filter(
        models.Framework.name == "CIS Controls", models.Framework.version == "v8"
    ).first()
    if not fw:
        fw = models.Framework(name="CIS Controls", version="v8")
        db.add(fw)
        db.flush()

    imported = 0
    skipped = 0
    for row in rows:
        if not row or not row[0]:
            continue
        control_id = str(row[0]).strip()
        title = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        description = str(row[2]).strip() if len(row) > 2 and row[2] else ""

        existing = db.query(models.Control).filter(
            models.Control.framework_id == fw.id,
            models.Control.control_id == control_id,
        ).first()
        if existing:
            skipped += 1
            continue

        db.add(models.Control(
            framework_id=fw.id,
            control_id=control_id,
            title=title,
            description=description,
            required_tags=[],
            optional_tags=[],
            evidence=[],
        ))
        imported += 1

    db.commit()
    return {"imported": imported, "skipped": skipped,
            "message": f"Imported {imported} controls, skipped {skipped} duplicates"}
