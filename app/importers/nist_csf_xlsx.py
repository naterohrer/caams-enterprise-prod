"""Import NIST Cybersecurity Framework (CSF) 2.0 from official XLSX workbook.

The official workbook is published by NIST at:
  https://www.nist.gov/cyberframework/csf-20-resources-and-overview

The spreadsheet structure uses merged cells for Function and Category rows,
with individual rows for each Subcategory.  Column layout (auto-detected):

  Function  |  Category               |  Subcategory  |  Description
  ──────────┼─────────────────────────┼───────────────┼──────────────────
  GOVERN    │  Organizational Context │  GV.OC-01     │  The ...
  (GV)      │  (GV.OC)               │  GV.OC-02     │  ...

Controls in the database are stored at the **Category** level (e.g. GV.OC)
with Subcategories (GV.OC-01, GV.OC-02 …) stored as sub_controls.
"""

import io
import re
from typing import Optional

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app import models


# ---------------------------------------------------------------------------
# Tag mappings: required / optional tags per CSF 2.0 category
# ---------------------------------------------------------------------------
_TAG_MAP: dict[str, dict] = {
    # ── GOVERN ──────────────────────────────────────────────────────────────
    "GV.OC": {
        "required": [],
        "optional": ["IAM"],
        "evidence": [
            "Mission statement and stakeholder communication records",
            "Board-approved organisational context documentation",
            "Legal and regulatory obligation register reviewed annually",
        ],
    },
    "GV.RM": {
        "required": ["vulnerability-scanning"],
        "optional": ["threat-intelligence"],
        "evidence": [
            "Risk management strategy document with executive sign-off",
            "Risk appetite and tolerance statements",
            "Vulnerability scan schedule and results supporting risk prioritisation",
        ],
    },
    "GV.RR": {
        "required": [],
        "optional": ["IAM"],
        "evidence": [
            "Cybersecurity roles and responsibilities matrix (RACI)",
            "Job descriptions with cybersecurity accountability clauses",
            "Delegation of authority records for security decisions",
        ],
    },
    "GV.PO": {
        "required": [],
        "optional": [],
        "evidence": [
            "Information security policy suite with annual review records",
            "Policy distribution and acknowledgement logs",
            "Policy exception management process documentation",
        ],
    },
    "GV.OV": {
        "required": ["security-monitoring"],
        "optional": ["SIEM", "log-management"],
        "evidence": [
            "Cybersecurity programme performance metrics and KPIs reported to leadership",
            "Security monitoring dashboard reviewed by executive/board",
            "Programme review meeting minutes with corrective action tracking",
        ],
    },
    "GV.SC": {
        "required": ["vendor-management"],
        "optional": ["third-party-risk"],
        "evidence": [
            "Supply chain risk management policy and third-party inventory",
            "Vendor security assessment questionnaires and results",
            "Contractual cybersecurity requirements in supplier agreements",
        ],
    },
    # ── IDENTIFY ────────────────────────────────────────────────────────────
    "ID.AM": {
        "required": ["asset-inventory", "configuration-management"],
        "optional": [],
        "evidence": [
            "Asset inventory (hardware, software, data) with assigned ownership",
            "Configuration management database (CMDB) or equivalent",
            "Software bill of materials (SBOM) for critical systems",
        ],
    },
    "ID.RA": {
        "required": ["vulnerability-scanning"],
        "optional": ["threat-intelligence", "penetration-testing"],
        "evidence": [
            "Risk assessment methodology documentation",
            "Vulnerability scan reports with risk-rated findings",
            "Threat intelligence feeds consumed and acted upon",
            "Penetration test reports with remediation tracking",
        ],
    },
    "ID.IM": {
        "required": [],
        "optional": ["vulnerability-scanning"],
        "evidence": [
            "Lessons-learned records from incidents and exercises",
            "Cybersecurity improvement plan with prioritised actions",
            "Benchmarking results against industry peers or standards",
        ],
    },
    # ── PROTECT ─────────────────────────────────────────────────────────────
    "PR.AA": {
        "required": ["IAM", "MFA", "access-control"],
        "optional": ["PAM", "SSO", "zero-trust"],
        "evidence": [
            "Identity and access management system user access reports",
            "MFA enforcement coverage report across all in-scope systems",
            "Periodic access recertification records",
            "Privileged access management session logs",
        ],
    },
    "PR.AT": {
        "required": ["security-awareness-training"],
        "optional": ["phishing-simulation"],
        "evidence": [
            "Security awareness training completion records for all personnel",
            "Phishing simulation campaign results and trend data",
            "Role-based cybersecurity training records for high-risk roles",
        ],
    },
    "PR.DS": {
        "required": ["encryption-at-rest", "data-classification"],
        "optional": ["DLP", "encryption-in-transit"],
        "evidence": [
            "Data classification policy and inventory of sensitive data stores",
            "Encryption-at-rest configuration evidence for critical data",
            "DLP policy configuration and alert logs",
            "Encryption-in-transit (TLS) scan results for external endpoints",
        ],
    },
    "PR.PS": {
        "required": ["configuration-management", "patch-management"],
        "optional": ["hardening", "vulnerability-scanning", "SAST"],
        "evidence": [
            "System hardening standards and compliance scan results",
            "Patch management records showing critical patch deployment timelines",
            "Configuration baseline change approval records",
            "SAST scan results integrated into CI/CD pipeline",
        ],
    },
    "PR.IR": {
        "required": ["backup", "network-segmentation"],
        "optional": ["disaster-recovery", "network-firewall"],
        "evidence": [
            "Network segmentation diagram with CDE/critical asset boundaries",
            "Backup configuration and schedule records",
            "Backup restoration test results within defined RTO/RPO",
            "Firewall rule sets with documented business justification",
        ],
    },
    # ── DETECT ──────────────────────────────────────────────────────────────
    "DE.CM": {
        "required": ["security-monitoring", "log-management"],
        "optional": ["SIEM", "EDR", "network-monitoring", "vulnerability-scanning"],
        "evidence": [
            "Security monitoring platform coverage report across critical assets",
            "Log management configuration showing source ingestion and retention",
            "SIEM alert rules and evidence of daily analyst review",
            "EDR deployment and detection coverage report",
        ],
    },
    "DE.AE": {
        "required": ["SIEM", "log-management"],
        "optional": ["SOAR", "threat-intelligence"],
        "evidence": [
            "Adverse event analysis procedures and escalation thresholds",
            "SIEM correlation rules for multi-source event analysis",
            "Threat intelligence integration records showing IOC consumption",
            "SOAR playbook execution logs for automated triage",
        ],
    },
    # ── RESPOND ─────────────────────────────────────────────────────────────
    "RS.MA": {
        "required": ["incident-response"],
        "optional": ["SOAR"],
        "evidence": [
            "Incident response plan with defined roles and escalation paths",
            "Incident ticket records showing classification and priority assignment",
            "IR tabletop exercise results and after-action reports",
        ],
    },
    "RS.AN": {
        "required": ["log-management", "SIEM"],
        "optional": ["forensics", "threat-intelligence"],
        "evidence": [
            "Incident analysis reports with root cause documentation",
            "Log correlation evidence used in incident investigation",
            "Forensic investigation reports for significant incidents",
        ],
    },
    "RS.CO": {
        "required": ["incident-response"],
        "optional": [],
        "evidence": [
            "Incident communication plan for internal and external stakeholders",
            "Regulatory breach notification records with timestamps",
            "IR status update communications for significant incidents",
        ],
    },
    "RS.MI": {
        "required": ["incident-response", "endpoint-protection"],
        "optional": ["EDR", "network-firewall"],
        "evidence": [
            "Incident containment and eradication records",
            "EDR response action logs for endpoint-based incidents",
            "Firewall block/quarantine rule changes made during incidents",
        ],
    },
    # ── RECOVER ─────────────────────────────────────────────────────────────
    "RC.RP": {
        "required": ["backup", "disaster-recovery"],
        "optional": ["business-continuity", "backup-testing"],
        "evidence": [
            "Recovery plan documentation with defined RTO/RPO targets",
            "Disaster recovery test results and lessons-learned records",
            "Backup restoration test records demonstrating successful recovery",
        ],
    },
    "RC.CO": {
        "required": [],
        "optional": [],
        "evidence": [
            "Post-incident stakeholder communication records",
            "Public or regulatory communications issued after significant incidents",
            "Lessons-learned sessions documented and distributed",
        ],
    },
}


def _extract_id(text: str, pattern: str) -> Optional[str]:
    """Extract a parenthesised identifier from a combined label, e.g.
    'Organizational Context (GV.OC)' → 'GV.OC'.
    """
    m = re.search(pattern, text)
    return m.group(1) if m else None


def _strip_id(text: str) -> str:
    """Remove a trailing parenthesised identifier from a label."""
    return re.sub(r"\s*\([^)]+\)\s*$", "", text).strip()


def parse_nist_csf_xlsx(
    xlsx_path,
    version: str = "2.0",
    sheet_name: Optional[str] = None,
) -> dict:
    """Parse an official NIST CSF 2.0 XLSX workbook into a CAAMS framework dict.

    Returns a dict with keys: name, version, controls.
    Each control corresponds to a CSF Category (e.g. GV.OC) and carries
    sub_controls for each Subcategory (e.g. GV.OC-01).

    The workbook uses merged cells for Function and Category — the parser
    tracks the last-seen non-empty value in those columns to handle this.
    """
    wb = load_workbook(xlsx_path if not isinstance(xlsx_path, (bytes, bytearray))
                       else io.BytesIO(xlsx_path),
                       read_only=False, data_only=True)

    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"
            )
        ws = wb[sheet_name]
    else:
        # Prefer a sheet whose name contains "core" (CSF naming convention)
        ws = next(
            (wb[n] for n in wb.sheetnames if "core" in n.lower()),
            wb.active,
        )

    rows = [
        tuple(cell.value for cell in row)
        for row in ws.iter_rows()
    ]
    if not rows:
        return {"name": "NIST Cybersecurity Framework", "version": version, "controls": []}

    # ── Detect header row ───────────────────────────────────────────────────
    HEADER_KEYWORDS = {"function", "category", "subcategory", "identifier", "csf"}
    header_row_idx = 0
    headers: list[str] = []
    for i, row in enumerate(rows[:6]):
        cells = [str(c or "").strip().lower() for c in row]
        if sum(1 for c in cells if any(k in c for k in HEADER_KEYWORDS)) >= 2:
            header_row_idx = i
            headers = cells
            break

    def _col(*candidates: str) -> int:
        for name in candidates:
            for i, h in enumerate(headers):
                if name in h:
                    return i
        return -1

    func_col  = _col("function")
    cat_col   = _col("category")
    sub_col   = _col("subcategory", "identifier", "id")
    desc_col  = _col("description", "subcategory description", "outcome")

    # Positional fallbacks for workbooks without headers
    if func_col == -1:
        func_col = 0
    if cat_col == -1:
        cat_col = 1
    if sub_col == -1:
        sub_col = 2
    if desc_col == -1:
        desc_col = 3

    # ── Parse rows ──────────────────────────────────────────────────────────
    # CSF pattern: "GV.OC-01" for subcategory, "GV.OC" for category ID
    SUB_RE  = re.compile(r"^[A-Z]{2}\.[A-Z]{2,3}-\d{2}$")
    CAT_RE  = re.compile(r"[A-Z]{2}\.[A-Z]{2,3}")   # e.g. GV.OC inside a label
    FUNC_RE = re.compile(r"\(([A-Z]{2})\)")          # e.g. (GV) inside a label

    controls_map:   dict[str, dict] = {}
    controls_order: list[str]       = []

    cur_func_label = ""
    cur_cat_id     = ""
    cur_cat_title  = ""

    for row in rows[header_row_idx + 1:]:
        if not row or all(c is None for c in row):
            continue

        def _cell(idx: int) -> str:
            return str(row[idx]).strip() if idx < len(row) and row[idx] is not None else ""

        func_raw = _cell(func_col)
        cat_raw  = _cell(cat_col)
        sub_raw  = _cell(sub_col)
        desc_raw = _cell(desc_col)

        # Update running Function label
        if func_raw and FUNC_RE.search(func_raw):
            cur_func_label = func_raw

        # Update running Category when we see a new category cell
        if cat_raw:
            m = CAT_RE.search(cat_raw)
            if m:
                cur_cat_id    = m.group(0)
                cur_cat_title = _strip_id(cat_raw)

        # A subcategory row must match the ID pattern (e.g. GV.OC-01)
        if not SUB_RE.match(sub_raw):
            continue
        if not cur_cat_id:
            continue

        # Ensure the parent category control exists
        if cur_cat_id not in controls_map:
            tags = _TAG_MAP.get(cur_cat_id, {})
            controls_map[cur_cat_id] = {
                "control_id":    cur_cat_id,
                "title":         cur_cat_title or cur_cat_id,
                "description":   f"CSF 2.0 category under the {cur_func_label} function.",
                "required_tags": tags.get("required", []),
                "optional_tags": tags.get("optional", []),
                "evidence":      tags.get("evidence", []),
                "sub_controls":  [],
            }
            controls_order.append(cur_cat_id)

        controls_map[cur_cat_id]["sub_controls"].append(
            {"id": sub_raw, "title": desc_raw, "description": ""}
        )

    return {
        "name":     "NIST Cybersecurity Framework",
        "version":  version,
        "controls": [controls_map[cid] for cid in controls_order],
    }


def import_nist_csf_xlsx(content: bytes, db: Session) -> dict:
    """Parse *content* and upsert controls into the database.

    - Creates the framework row if it does not exist.
    - Adds new controls; skips controls whose control_id already exists.
    - Updates sub_controls on existing controls (backfill).
    Returns a summary dict with keys: imported, skipped, message.
    """
    framework = parse_nist_csf_xlsx(content)
    version   = framework["version"]

    fw = db.query(models.Framework).filter(
        models.Framework.name    == "NIST Cybersecurity Framework",
        models.Framework.version == version,
    ).first()
    if not fw:
        fw = models.Framework(
            name    = "NIST Cybersecurity Framework",
            version = version,
        )
        db.add(fw)
        db.flush()

    imported = 0
    skipped  = 0
    for ctrl in framework["controls"]:
        existing = db.query(models.Control).filter(
            models.Control.framework_id == fw.id,
            models.Control.control_id   == ctrl["control_id"],
        ).first()

        if existing:
            # Backfill sub_controls if they were empty
            if not existing.sub_controls and ctrl["sub_controls"]:
                existing.sub_controls = ctrl["sub_controls"]
            skipped += 1
            continue

        db.add(models.Control(
            framework_id   = fw.id,
            control_id     = ctrl["control_id"],
            title          = ctrl["title"],
            description    = ctrl["description"],
            required_tags  = ctrl["required_tags"],
            optional_tags  = ctrl["optional_tags"],
            evidence       = ctrl["evidence"],
            sub_controls   = ctrl["sub_controls"],
        ))
        imported += 1

    db.commit()
    return {
        "imported": imported,
        "skipped":  skipped,
        "message":  f"NIST CSF {version}: {imported} categories imported, {skipped} skipped.",
    }
